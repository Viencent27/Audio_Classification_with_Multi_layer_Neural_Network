import os
import glob
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from keras.utils import to_categorical
from keras.models import save_model, load_model
from sklearn.model_selection import train_test_split
from Machine_Learning import create_model, train_model, evaluate_model
from GUI import deploy_web_app
from Audio_Preprocessing import preprocess_audio

def load_data(path, target_sample_rate):
  """
  Memuat data audio dari file .wav sesuai dengan path yang diberikan.
  """
  audio_files = glob.glob(path)
  X = []
  y = []

  for audio_file in audio_files:
    mfccs_mean, mfccs = preprocess_audio(audio_file, target_sample_rate)
    X.append(mfccs_mean)
    y.append(0 if 'happy' in audio_file else 1)

  X = np.array(X)
  y = np.array(y)

  return X, y

def main(target_sample_rate):
  """
  Fungsi utama yang menjalankan alur aplikasi.
  """
  project_directory = r"C:\Users\Lenovo\Documents\Bljr Coding\Python\PPDM\Final Project"
  audiosamples_directory = os.path.join(project_directory, "audiosamples")

  X, y = load_data(os.path.join(audiosamples_directory, "*.wav"), target_sample_rate)

  num_classes = 2  # Jumlah kelas (positive sentiment dan negative sentiment)
  input_shape = (13,)  # Bentuk data input (fitur MFCC)

  num_samples = len(X)
  num_test_samples = int(0.2*num_samples)

  if num_samples > 0 and num_test_samples > 0:
    test_size = num_test_samples / num_samples
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size)
    y_train = to_categorical(y_train, num_classes)
    y_test = to_categorical(y_test, num_classes)

    learning_rates = [0.01, 0.001, 0.0001]
    num_epochs_list = [50, 100, 200]
    hidden_layers_list = [1, 2, 3]
    hidden_neurons_list = [32, 64, 128]
    activations = ['sigmoid', 'softmax', 'tanh']

    if os.path.exists('best_model.h5'):
      # Load the saved model
      best_model = load_model('best_model.h5')
      deploy_web_app(best_model, X_test, y_test, target_sample_rate)
    
    else:
      best_accuracy = 0
      best_model = None
      hyper_combination_data = []

      for lr in learning_rates:
        for num_epochs in num_epochs_list:
          for num_hidden_layers in hidden_layers_list:
            for num_hidden_neurons in hidden_neurons_list:
              for activation in activations:
                model = create_model(input_shape, num_classes, num_hidden_layers, num_hidden_neurons, activation)
                model = train_model(model, X_train, y_train, lr, num_epochs)
                accuracy, precision, recall, f1 = evaluate_model(model, X_test, y_test)
                combination_row_data = [lr, num_epochs, num_hidden_layers, num_hidden_neurons, activation,
                                        accuracy, precision, recall, f1]
                hyper_combination_data.append(combination_row_data)

                if accuracy > best_accuracy:
                  best_accuracy = accuracy
                  best_model = model

                print(f"Learning Rate: {lr}, Epochs: {num_epochs}, Hidden Layers: {num_hidden_layers}, Hidden Neurons: {num_hidden_neurons}, Activation: {activation}")
                print(f"Accuracy: {accuracy}, Precision: {precision}, Recall: {recall}, F1-Score: {f1}")
                print("=========================================")
    
      if best_model is not None:
        # Save the best model to a file
        save_model(best_model, 'best_model.h5')

        # Load the saved model
        loaded_model = load_model('best_model.h5')

        # Membuat objek Workbook
        workbook = Workbook()
        sheet = workbook.active

        # Menulis data ke file excel
        sheet['A1'] = 'Learning Rate'
        sheet['B1'] = 'Epochs'
        sheet['C1'] = 'Hidden Layers'
        sheet['D1'] = 'Hidden Neurons'
        sheet['E1'] = 'Activation'
        sheet['F1'] = 'Accuracy'
        sheet['G1'] = 'Precision'
        sheet['H1'] = 'Recall'
        sheet['I1'] = 'F1-Score'

        for row_index, row_data in enumerate(hyper_combination_data, start=2):
          sheet.append(row_data)
          if row_data[5] == best_accuracy: # Memberikan style ke data dengan akurasi paling tinggi
            for col_index in range(1, sheet.max_column + 1):
              cell = sheet.cell(row=row_index, column=col_index)
              cell.font = Font(color="FF0000")

        # Menyimpan file Excel
        workbook.save(filename='Data Kombinasi Hyper Parameter.xlsx')
        
        deploy_web_app(loaded_model, X_test, y_test, target_sample_rate)
      else:
        print("Error: Model terbaik tidak ditemukan.")

  else:
    print("Error: Data tidak cukup untuk dibagi menjadi data training dan data testing.")

# Contoh penggunaan
if __name__ == '__main__':
  main(44100)  # Set the target sampling rate to 44100 Hz